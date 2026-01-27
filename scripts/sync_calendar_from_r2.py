#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import os
import json
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

import boto3
import requests
import openpyxl

try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None


# ----------------------------
# Helpers
# ----------------------------

def die(msg: str, code: int = 1) -> None:
    print(f"[ERRO] {msg}")
    raise SystemExit(code)


def env_required(name: str) -> str:
    v = os.getenv(name)
    if not v:
        die(f"Variável de ambiente obrigatória não definida: {name}")
    return v


def parse_letivo(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in {"sim", "true", "1", "yes", "y", "letivo"}


def parse_date_cell(value: Any) -> Optional[date]:
    if value is None:
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


# ----------------------------
# R2
# ----------------------------

def r2_client():
    return boto3.client(
        "s3",
        endpoint_url=env_required("R2_ENDPOINT"),
        aws_access_key_id=env_required("R2_ACCESS_KEY_ID"),
        aws_secret_access_key=env_required("R2_SECRET_ACCESS_KEY"),
        region_name="auto",
    )


def download_text_from_r2(bucket: str, key: str) -> str:
    s3 = r2_client()
    obj = s3.get_object(Bucket=bucket, Key=key)
    return obj["Body"].read().decode("utf-8")


def download_file_from_r2(bucket: str, key: str, out_path: str) -> None:
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    s3 = r2_client()
    s3.download_file(bucket, key, out_path)
    print(f"OK download: {key} -> {out_path}")


# ----------------------------
# Excel
# ----------------------------

def read_export_sheet(xlsx_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        die(f"Aba '{sheet_name}' não encontrada. Abas: {wb.sheetnames}")

    ws = wb[sheet_name]
    print(f"Aba selecionada: '{sheet_name}'")

    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col

    if "data" not in headers or "letivo" not in headers:
        die("A aba export precisa ter colunas: data | letivo")

    items: List[Dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        d = parse_date_cell(ws.cell(row=row, column=headers["data"]).value)
        if not d:
            continue

        letivo = ws.cell(row=row, column=headers["letivo"]).value
        items.append({
            "day": d.isoformat(),
            "is_school_day": parse_letivo(letivo),
        })

    print(f"Linhas lidas: {len(items)}")
    print("Amostra:")
    print(json.dumps(items[:5], indent=2, ensure_ascii=False))
    return items


# ----------------------------
# API
# ----------------------------

def api_login(api_base_url: str, username: str) -> str:
    url = api_base_url.rstrip("/") + "/auth/login"
    resp = requests.post(url, json={"username": username}, timeout=30)

    if resp.status_code != 200:
        die(f"Login falhou ({resp.status_code}): {resp.text}")

    token = resp.json().get("access_token")
    if not token:
        die("Login OK, mas access_token ausente")

    print("OK login automático. Token recebido.")
    return token


def post_calendar_import(api_base_url: str, token: str, rows: List[Dict[str, Any]]) -> None:
    batch_size = int(os.getenv("BATCH_SIZE", "50"))
    url = api_base_url.rstrip("/") + "/calendar/import"

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "accept": "application/json",
    }

    total = len(rows)
    batches = [rows[i:i + batch_size] for i in range(0, total, batch_size)]

    for i, batch in enumerate(batches, start=1):
        print(f"→ POST lote {i}/{len(batches)} (itens {len(batch)})")
        resp = requests.post(url, json=batch, headers=headers, timeout=120)

        if resp.status_code != 200:
            die(f"Erro no lote {i}: {resp.status_code} - {resp.text}")

        print("API status:", resp.status_code)

    print("✅ Importação concluída com sucesso (todos os lotes).")


# ----------------------------
# Main
# ----------------------------

def main():
    if load_dotenv:
        load_dotenv()

    bucket = env_required("R2_BUCKET")
    index_key = env_required("R2_INDEX_KEY")
    api_base_url = env_required("API_BASE_URL")
    login_username = os.getenv("LOGIN_USERNAME", "paulo")
    sheet_name = os.getenv("SHEET_NAME", "export")

    # 1. Lê index.json
    index_raw = download_text_from_r2(bucket, index_key)
    index = json.loads(index_raw)

    current_key = index.get("current_key")
    if not current_key:
        die("index.json não possui current_key")

    # 2. Baixa XLSX atual
    out_path = "tmp/calendar_current.xlsx"
    download_file_from_r2(bucket, current_key, out_path)

    # 3. Lê planilha
    rows = read_export_sheet(out_path, sheet_name)

    # 4. Limite opcional
    limit = os.getenv("LIMIT_DAYS")
    if limit:
        rows = rows[:int(limit)]
        print(f"LIMIT_DAYS aplicado: {limit}")

    # 5. Login + import
    token = api_login(api_base_url, login_username)
    post_calendar_import(api_base_url, token, rows)


if __name__ == "__main__":
    main()